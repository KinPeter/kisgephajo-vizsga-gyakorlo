import openpyxl
import json
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("hsz2024.xlsx")
ws = wb.active

if not ws:
    raise ValueError("No active worksheet found in the workbook.")

data_prep = []
for row in ws.iter_rows():
    row_dict = {}
    for cell in row:
        if not cell.column:
            continue
        col_letter = get_column_letter(cell.column)

        match col_letter:
            case "A":
                row_dict["question"] = cell.value
            case "B":
                row_dict["answer_1"] = cell.value
            case "C":
                row_dict["is_answer_1_correct"] = (
                    str(cell.value).lower().strip() == "igen"
                )
            case "D":
                row_dict["answer_2"] = cell.value
            case "E":
                row_dict["is_answer_2_correct"] = (
                    str(cell.value).lower().strip() == "igen"
                )
            case "F":
                row_dict["answer_3"] = cell.value
            case "G":
                row_dict["is_answer_3_correct"] = (
                    str(cell.value).lower().strip() == "igen"
                )
            case "H":
                row_dict["image"] = cell.value
            case "J":
                row_dict["id"] = cell.value

    data_prep.append(row_dict)


data = []
for item in data_prep[1:]:  # Skip header row
    # Find which answer is correct
    answers = [
        (item.get("answer_1"), item.get("is_answer_1_correct")),
        (item.get("answer_2"), item.get("is_answer_2_correct")),
        (item.get("answer_3"), item.get("is_answer_3_correct")),
    ]
    correct = None
    incorrect = []
    for ans, is_corr in answers:
        if is_corr:
            correct = ans
        else:
            incorrect.append(ans)
    # Pad incorrect answers to always have 2 (with None if missing)
    while len(incorrect) < 2:
        incorrect.append(None)
    data.append(
        {
            "question": item.get("question"),
            "correct": correct,
            "incorrect1": incorrect[0],
            "incorrect2": incorrect[1],
            "image": item.get("image"),
            "id": item.get("id"),
        }
    )

# Export to JSON
with open("hsz_output.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
