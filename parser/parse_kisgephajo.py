import openpyxl
import json
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("kisgephajo2022.xlsx")
ws = wb.active

if not ws:
    raise ValueError("No active worksheet found in the workbook.")

data = []
for row in ws.iter_rows():
    row_dict = {}
    for cell in row:
        if not cell.column:
            continue
        col_letter = get_column_letter(cell.column)

        match col_letter:
            case "A":
                row_dict["id"] = int(str(cell.value)) if cell.value != "Szám" else None
            case "B":
                row_dict["question"] = cell.value
            case "C":
                row_dict["correct"] = cell.value
            case "D":
                row_dict["incorrect1"] = cell.value
            case "E":
                row_dict["incorrect2"] = cell.value
            case "F":
                row_dict["image"] = (
                    str(cell.value).replace(" ", "_") if cell.value else None
                )

    data.append(row_dict)

# Export to JSON
with open("hi_output.json", "w", encoding="utf-8") as f:
    json.dump(data[1:], f, ensure_ascii=False, indent=2)
